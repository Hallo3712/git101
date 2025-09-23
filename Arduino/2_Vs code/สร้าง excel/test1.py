from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

# สร้าง workbook และ sheet
wb = Workbook()

# กำหนดความยาวแถวที่ต้องการลงสีและใส่ dropdown
max_row_fill = 100

# --- สร้าง sheet "สรุป" ---
ws_summary = wb.active
ws_summary.title = "สรุป"

# กำหนดหัวตาราง และเติมสี
header_fill_summary = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
cell_fill_summary = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
border_side = Side(border_style="thin", color="000000")
border_all = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

summary_headers = ["หมวดหมู่", "ยอดรวมรายรับ", "ยอดรวมรายจ่าย"]
ws_summary.append(summary_headers)

for col_num in range(1, len(summary_headers) + 1):
    cell = ws_summary.cell(row=1, column=col_num)
    cell.fill = header_fill_summary
    cell.border = border_all
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center")

# เติมสีและเส้นกรอบให้แถวที่เหลือ (จนถึง max_row_fill)
for row in range(2, max_row_fill + 1):
    for col in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=row, column=col)
        cell.fill = cell_fill_summary
        cell.border = border_all


# --- สร้าง sheet "รายวัน" ---
ws_daily = wb.create_sheet("รายวัน")

daily_headers = ["วันที่", "บัญชี", "จำนวนเงิน", "หมวดหมู่"]
ws_daily.append(daily_headers)

header_fill_daily = PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid")
cell_fill_daily = PatternFill(start_color="E4D9F2", end_color="E4D9F2", fill_type="solid")

for col_num in range(1, len(daily_headers) + 1):
    cell = ws_daily.cell(row=1, column=col_num)
    cell.fill = header_fill_daily
    cell.border = border_all
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center")

for row in range(2, max_row_fill + 1):
    for col in range(1, len(daily_headers) + 1):
        cell = ws_daily.cell(row=row, column=col)
        cell.fill = cell_fill_daily
        cell.border = border_all

# --- สร้าง sheet "ตัวเลือก" ---
ws_options = wb.create_sheet("ตัวเลือก")

options_headers = ["บัญชี", "ยอดคงเหลือ", "หมวดหมู่"]
ws_options.append(options_headers)

header_fill_options = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
cell_fill_options = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")

for col_num in range(1, len(options_headers) + 1):
    cell = ws_options.cell(row=1, column=col_num)
    cell.fill = header_fill_options
    cell.border = border_all
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

for row in range(2, max_row_fill + 1):
    for col in range(1, len(options_headers) + 1):
        cell = ws_options.cell(row=row, column=col)
        cell.fill = cell_fill_options
        cell.border = border_all

# --- สร้าง dropdown data validation ใน "รายวัน" ---
# dropdown บัญชี อ้างอิงจากตัวเลือก!A2:A100
dv_account = DataValidation(type="list", formula1='ตัวเลือก!$A$2:$A$100', allow_blank=True)
ws_daily.add_data_validation(dv_account)
for row in range(2, max_row_fill + 1):
    dv_account.add(ws_daily.cell(row=row, column=2))  # คอลัมน์ B = บัญชี

# dropdown หมวดหมู่ อ้างอิงจากตัวเลือก!C2:C100
dv_category = DataValidation(type="list", formula1='ตัวเลือก!$C$2:$C$100', allow_blank=True)
ws_daily.add_data_validation(dv_category)
for row in range(2, max_row_fill + 1):
    dv_category.add(ws_daily.cell(row=row, column=4))  # คอลัมน์ D = หมวดหมู่

# --- บันทึกไฟล์ ---
filename = "รายรับรายจ่าย.xlsx"
wb.save(filename)
print(f"สร้างไฟล์ {filename} เรียบร้อยแล้ว")
